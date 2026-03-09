#!/usr/bin/env python3
"""
Prepare Sinhala TTS training data for Piper.

Reads TextData.xlsx and wav files from the input directory, then:
  - Creates metadata.csv in pipe-delimited LJSpeech format: id|text
  - Resamples all wav files from 44.1kHz to 22050Hz using sox
  - Validates that every wav referenced in the spreadsheet exists
  - Outputs to prepared_data/wav/ and prepared_data/metadata.csv
"""

import argparse
import os
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(description="Prepare Sinhala TTS data for Piper training")
    parser.add_argument(
        "--input-dir",
        type=str,
        default="SinhalaTTSData",
        help="Path to input directory containing TextData.xlsx and wavs/ (default: SinhalaTTSData)",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default="prepared_data",
        help="Path to output directory (default: prepared_data)",
    )
    parser.add_argument(
        "--sample-rate",
        type=int,
        default=22050,
        help="Target sample rate in Hz (default: 22050)",
    )
    return parser.parse_args()


def read_xlsx(xlsx_path: Path) -> list[tuple[str, str]]:
    """Read TextData.xlsx and return list of (filename_stem, text) tuples."""
    print(f"Reading {xlsx_path} ...")
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    entries = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if row is None or len(row) < 2:
            continue
        col_a, col_b = row[0], row[1]
        if col_a is None or col_b is None:
            continue

        # Determine which column has the filename and which has the text.
        # The filename column contains values like "a000001" or "a000001.wav".
        # The text column contains Sinhala script.
        filename_val = str(col_a).strip()
        text_val = str(col_b).strip()

        # Skip header rows if present
        if row_idx == 1 and ("file" in filename_val.lower() or "name" in filename_val.lower()):
            continue

        # Strip .wav extension if present
        if filename_val.lower().endswith(".wav"):
            filename_val = filename_val[:-4]

        if not filename_val or not text_val:
            continue

        entries.append((filename_val, text_val))

    wb.close()
    print(f"  Found {len(entries)} entries in spreadsheet")
    return entries


def validate_data(entries: list[tuple[str, str]], wav_dir: Path) -> list[tuple[str, str]]:
    """Validate that wav files exist for all spreadsheet entries. Return valid entries."""
    available_wavs = {p.stem for p in wav_dir.glob("*.wav")}
    referenced_stems = {stem for stem, _ in entries}

    missing_wavs = referenced_stems - available_wavs
    extra_wavs = available_wavs - referenced_stems

    if missing_wavs:
        print(f"\n  WARNING: {len(missing_wavs)} files in spreadsheet but missing from wavs/:")
        for name in sorted(missing_wavs)[:10]:
            print(f"    - {name}.wav")
        if len(missing_wavs) > 10:
            print(f"    ... and {len(missing_wavs) - 10} more")

    if extra_wavs:
        print(f"\n  INFO: {len(extra_wavs)} wav files not referenced in spreadsheet:")
        for name in sorted(extra_wavs)[:10]:
            print(f"    - {name}.wav")
        if len(extra_wavs) > 10:
            print(f"    ... and {len(extra_wavs) - 10} more")

    valid_entries = [(stem, text) for stem, text in entries if stem in available_wavs]
    print(f"\n  Valid entries with matching wav files: {len(valid_entries)}")
    return valid_entries


def resample_wav(src: Path, dst: Path, target_sr: int):
    """Resample a wav file using sox with -v 0.95 to avoid clipping."""
    cmd = [
        "sox", str(src),
        "-r", str(target_sr),
        "-c", "1",         # mono
        "-b", "16",        # 16-bit
        str(dst),
        "vol", "0.95",     # slight volume reduction to avoid clipping
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"  ERROR resampling {src.name}: {result.stderr.strip()}")
        return False
    return True


def main():
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    target_sr = args.sample_rate

    # Validate input directory
    xlsx_path = input_dir / "TextData.xlsx"
    wav_dir = input_dir / "wavs"

    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found")
        sys.exit(1)
    if not wav_dir.exists():
        print(f"ERROR: {wav_dir} not found")
        sys.exit(1)

    # Check sox is available
    try:
        subprocess.run(["sox", "--version"], capture_output=True, check=True)
    except FileNotFoundError:
        print("ERROR: sox is not installed. Install with: brew install sox (macOS) or apt install sox (Linux)")
        sys.exit(1)

    # Read and validate data
    entries = read_xlsx(xlsx_path)
    valid_entries = validate_data(entries, wav_dir)

    if not valid_entries:
        print("ERROR: No valid entries found. Exiting.")
        sys.exit(1)

    # Create output directories
    out_wav_dir = output_dir / "wav"
    out_wav_dir.mkdir(parents=True, exist_ok=True)
    print(f"\nOutput directory: {output_dir}")

    # Write metadata.csv
    metadata_path = output_dir / "metadata.csv"
    with open(metadata_path, "w", encoding="utf-8") as f:
        for stem, text in valid_entries:
            f.write(f"{stem}|{text}\n")
    print(f"Wrote {len(valid_entries)} entries to {metadata_path}")

    # Resample wav files
    print(f"\nResampling {len(valid_entries)} wav files to {target_sr}Hz ...")
    success_count = 0
    fail_count = 0
    for i, (stem, _) in enumerate(valid_entries):
        src = wav_dir / f"{stem}.wav"
        dst = out_wav_dir / f"{stem}.wav"

        if resample_wav(src, dst, target_sr):
            success_count += 1
        else:
            fail_count += 1

        # Progress every 100 files
        done = i + 1
        if done % 100 == 0 or done == len(valid_entries):
            pct = done / len(valid_entries) * 100
            print(f"  [{done}/{len(valid_entries)}] {pct:.1f}% complete")

    print(f"\nDone! Resampled {success_count} files, {fail_count} failures.")
    print(f"Output:")
    print(f"  Metadata: {metadata_path}")
    print(f"  Wav files: {out_wav_dir}/")


if __name__ == "__main__":
    main()
